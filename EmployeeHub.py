import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import Calendar
import os
import sqlite3
from datetime import timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import Font

class SummaryPopup:
    def __init__(self, parent):
        self.parent = parent
        self.popup = tk.Toplevel(parent)
        self.popup.title("Summary")
        self.create_widgets()

    def create_widgets(self):
        self.filter_frame = tk.Frame(self.popup)
        self.filter_frame.pack(padx=10, pady=10)       
        filter_label_employee_name = tk.Label(self.filter_frame, text="Employee Name:")
        filter_label_employee_name.grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.filter_entry_employee_name = tk.Entry(self.filter_frame)
        self.filter_entry_employee_name.grid(row=0, column=1, padx=5, pady=5)       
        filter_label_position = tk.Label(self.filter_frame, text="Position:")
        filter_label_position.grid(row=0, column=2, padx=5, pady=5, sticky="e")
        self.filter_entry_position = tk.Entry(self.filter_frame)
        self.filter_entry_position.grid(row=0, column=3, padx=5, pady=5)      
        filter_label_month = tk.Label(self.filter_frame, text="Month:")
        filter_label_month.grid(row=0, column=4, padx=5, pady=5, sticky="e")
        self.filter_entry_month = tk.Entry(self.filter_frame)
        self.filter_entry_month.grid(row=0, column=5, padx=5, pady=5)       
        filter_label_year = tk.Label(self.filter_frame, text="Year:")
        filter_label_year.grid(row=0, column=6, padx=5, pady=5, sticky="e")
        self.filter_entry_year = tk.Entry(self.filter_frame)
        self.filter_entry_year.grid(row=0, column=7, padx=5, pady=5)       
        self.filter_button = tk.Button(self.filter_frame, text="Filter", command=self.filter_summary)
        self.filter_button.grid(row=0, column=8, padx=5, pady=5)
        export_button = tk.Button(self.filter_frame, text="Export to Excel", command=self.export_to_excel)
        export_button.grid(row=0, column=9, padx=5, pady=5)
        self.tree = ttk.Treeview(self.popup, columns=("Employee_Name", "Position", "Month", "Year", "Total_Work_Time", "Total_Break_Time", "Net_Work_Time", "Total_Annual_Leave", "Total_Sick_Leave"), show="headings")
        self.tree.heading("Employee_Name", text="Name", anchor="w")
        self.tree.heading("Position", text="Position", anchor="w")
        self.tree.heading("Month", text="Month", anchor="w")
        self.tree.heading("Year", text="Year", anchor="w")
        self.tree.heading("Total_Work_Time", text="Work Time", anchor="w")
        self.tree.heading("Total_Break_Time", text="Break Time", anchor="w")
        self.tree.heading("Net_Work_Time", text="Work(Net)", anchor="w")
        self.tree.heading("Total_Annual_Leave", text="Annual Leave", anchor="w")
        self.tree.heading("Total_Sick_Leave", text="Sick Leave", anchor="w")
        self.tree.column("Employee_Name", width=100)
        self.tree.column("Position", width=100)
        self.tree.column("Month", width=50)
        self.tree.column("Year", width=50)
        self.tree.column("Total_Work_Time", width=120)
        self.tree.column("Total_Break_Time", width=120)
        self.tree.column("Net_Work_Time", width=120)     
        self.tree.column("Total_Annual_Leave", width=120)
        self.tree.column("Total_Sick_Leave", width=120)
        self.tree.pack(padx=10, pady=10)
        self.load_summary()

    def load_summary(self):
        try:
            conn = sqlite3.connect("employees.db")
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM summary")
            summary_data = cursor.fetchall()
            for data in summary_data:
                data = list(data)
                data[4] = self.seconds_to_hours_minutes(data[4])
                data[5] = self.seconds_to_hours_minutes(data[5])
                data[6] = self.seconds_to_hours_minutes(data[6])
                total_annual_leave = data[9]
                total_sick_leave = data[10]
                data.append(total_annual_leave)
                data.append(total_sick_leave)
                
                self.tree.insert("", "end", values=data[:11])
            conn.close()
        except Exception as e:
            print("An error occurred while loading the summary:", e)

    def seconds_to_hours_minutes(self, seconds):
        td = timedelta(seconds=seconds)
        hours, remainder = divmod(td.seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}"
    
    def filter_summary(self):
        employee_name = self.filter_entry_employee_name.get()
        position = self.filter_entry_position.get()
        month = self.filter_entry_month.get()
        year = self.filter_entry_year.get()
        try:
            conn = sqlite3.connect("employees.db")
            cursor = conn.cursor()
            summary_query = "SELECT * FROM summary WHERE 1=1"
            if employee_name:
                summary_query += f" AND Employee_Name LIKE '%{employee_name}%'"
            if position:
                summary_query += f" AND Position LIKE '%{position}%'"
            if month:
                summary_query += f" AND Month = '{month}'"
            if year:
                summary_query += f" AND Year = '{year}'"
            cursor.execute(summary_query)
            summary_data = cursor.fetchall()
            conn.close()
            self.tree.delete(*self.tree.get_children())
            for data in summary_data:
                data = list(data)
                data[4] = self.seconds_to_hours_minutes(data[4])
                data[5] = self.seconds_to_hours_minutes(data[5])
                data[6] = self.seconds_to_hours_minutes(data[6])
                data[9] = self.seconds_to_hours_minutes(data[9])
                data[10] = self.seconds_to_hours_minutes(data[10])
                self.tree.insert("", "end", values=data)
        except Exception as e:
            print("An error occurred while loading the summary:", e)

    def export_to_excel(self):
        try:
            conn = sqlite3.connect("employees.db")
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM summary")
            summary_data = cursor.fetchall()
            conn.close()
            filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
            if filename:
                workbook = Workbook()
                worksheet = workbook.active
                bold = Font(bold=True)
                headings = ["Employee Name", "Position", "Month", "Year", "Total Work Time", "Total Break Time", "Net Work Time"]
                for i, heading in enumerate(headings):
                    worksheet.cell(row=1, column=i+1, value=heading).font = bold
                for row_num, row_data in enumerate(summary_data, start=2):
                    for col_num, cell_data in enumerate(row_data, start=1):
                        worksheet.cell(row=row_num, column=col_num, value=cell_data)
                workbook.save(filename)
        except Exception as e:
            print("An error occurred while exporting to Excel:", e)


class AddWorkHoursPopup:
    def __init__(self, parent, callback, employee_id=None, employee_name=None):
        self.parent = parent
        self.employee_id = employee_id
        self.employee_name = employee_name
        self.callback = callback
        self.popup = tk.Toplevel(parent)
        self.popup.title("Add Work Hours")
        self.create_widgets()

    def create_widgets(self):
        label_date = tk.Label(self.popup, text="Date (DD-MM-YYYY):")
        label_date.grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.calendar = Calendar(self.popup, selectmode="day", date_pattern="dd-MM-yyyy")
        self.calendar.grid(row=0, column=1, padx=5, pady=5)
        times = ["Start Time", "End Time", "Break Start Time", "Break End Time"]
        rows = [1, 2, 3, 4]
        self.time_spinboxes = []
        for idx, time in enumerate(times):
            label = tk.Label(self.popup, text=f"{time} (HH:MM):")
            label.grid(row=rows[idx], column=0, padx=5, pady=5, sticky="e")
            hour_spinbox, minute_spinbox = self.create_time_spinboxes(self.popup, row=rows[idx], column=1)
            self.time_spinboxes.append((hour_spinbox, minute_spinbox))
        save_button = tk.Button(self.popup, text="Save", command=self.validate_and_save_work_hours)
        save_button.grid(row=6, columnspan=2, padx=5, pady=10)

    def create_time_spinboxes(self, parent, row, column):
        spinbox_hour = tk.Spinbox(parent, from_=0, to=23, width=5)
        spinbox_hour.grid(row=row, column=column, padx=(0, 5), pady=5)
        spinbox_minute = tk.Spinbox(parent, from_=0, to=59, width=5)
        spinbox_minute.grid(row=row, column=column + 1, padx=(0, 5), pady=5)
        return spinbox_hour, spinbox_minute
    
    def is_valid_time_format(self, time_str):
        try:
            hour, minute = time_str.split(":")
            if 0 <= int(hour) <= 23 and 0 <= int(minute) <= 59:
                return True
            else:
                return False
        except ValueError:
            return False

    def are_working_hours_valid(self, start_time, end_time, break_start, break_end):
        try:
            if not self.is_valid_order(start_time, end_time):
                print("Start time is not before end time")
                return False
            if break_start and (not self.is_valid_order(start_time, break_start) or not self.is_valid_order(break_start, end_time)):
                print("Break start is not after start time or not before end time")
                return False
            if break_end and (not self.is_valid_order(start_time, break_end) or not self.is_valid_order(break_end, end_time)):
                print("Break end is not after start time or not before end time")
                return False
            if break_start and break_end:
                if not self.is_valid_order(break_start, break_end):
                    print("Break end time is not after break start time")
                    return False                            
            return True
        except ValueError:
            return False

    def is_valid_order(self, time1, time2):
        if time1 is None or time2 is None:
            return True
        hour1, minute1 = map(int, time1.split(":"))
        hour2, minute2 = map(int, time2.split(":"))
        if hour1 < hour2 or (hour1 == hour2 and minute1 <= minute2):
            return True
        print(time1)
        print(time2)
        return False

    def validate_and_save_work_hours(self):
        date = self.calendar.get_date()
        start_time = self.get_time_from_spinboxes(1)
        end_time = self.get_time_from_spinboxes(2)
        break_start = self.get_time_from_spinboxes(3)
        break_end = self.get_time_from_spinboxes(4)
        if start_time == "00:00":
            start_time = None
        if end_time == "00:00":
            end_time = None
        if break_start == "00:00":
            break_start = None
        if break_end == "00:00":
            break_end = None
        if not (date and start_time and end_time):
            messagebox.showerror("Error", "Date, Start Time, and End Time fields are mandatory.")
            return
        if not all(self.is_valid_time_format(time_str) for time_str in [start_time, end_time]):
            messagebox.showerror("Error", "Invalid time format. Please use HH:MM format.")
            return
        if not self.are_working_hours_valid(start_time, end_time, break_start, break_end):
            messagebox.showerror("Error", "Invalid work hours. Please check your inputs.")
            return
        if break_start and not break_end:
            messagebox.showerror("Error", "Break end time is missing. Please provide break end time.")
            return
        if not break_start:
            break_start = None
        if break_start and not break_end:
            messagebox.showerror("Error", "Break end time is missing. Please provide break end time.")
            return
        if not break_start and break_end:
            messagebox.showerror("Error", "Break start time is missing. Please provide break start time.")
            return
        self.save_work_hours(date, start_time, end_time, break_start, break_end)

    def save_work_hours(self, date, start_time, end_time, break_start, break_end):
        if not (date and start_time and end_time):
            messagebox.showerror("Error", "Date, Start Time, and End Time fields are mandatory.")
            return
        if not all(self.is_valid_time_format(time_str) for time_str in [start_time, end_time]):
            messagebox.showerror("Error", "Invalid time format. Please use HH:MM format.")
            return
        if not self.are_working_hours_valid(start_time, end_time, break_start, break_end):
            messagebox.showerror("Error", "Invalid work hours. Please check your inputs.")
            return
        try:
            conn = sqlite3.connect("employees.db")
            cursor = conn.cursor()
            cursor.execute("INSERT INTO work_hours (Employee_ID, Date, Start_Time, End_Time, Break_Start, Break_End) VALUES (?, ?, ?, ?, ?, ?)",
                            (self.employee_id, date, start_time, end_time, break_start or None, break_end or None))
            conn.commit()
            conn.close()
            self.popup.destroy()
            work_hours = {"Date": date, "Start_Time": start_time, "End_Time": end_time, "Break_Start": break_start, "Break_End": break_end}
            self.callback(self.employee_id, self.employee_name, work_hours)
            print("Work hours saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving work hours: {e}")
  
    def get_time_from_spinboxes(self, row):
        spinbox_hour, spinbox_minute = self.time_spinboxes[row - 1]
        hour = spinbox_hour.get()
        minute = spinbox_minute.get()
        if not hour:
            return None
        if not minute:
            minute = "00"
        return f"{int(hour):02d}:{int(minute):02d}"
    
class AddLeaveRecordPopup:
    def __init__(self, parent, employee_id=None, employee_name=None):
        self.parent = parent
        self.employee_id = employee_id
        self.employee_name = employee_name
        self.popup = tk.Toplevel(parent)
        self.popup.title("Add Leave Record")
        self.create_widgets()

    def create_widgets(self):
        label_employee_name = tk.Label(self.popup, text=f"Employee: {self.employee_name}")
        label_employee_name.grid(row=0, columnspan=2, padx=5, pady=5)        
        label_leave_type = tk.Label(self.popup, text="Leave Type:")
        label_leave_type.grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.leave_type_combo = ttk.Combobox(self.popup, values=["Sick leave", "Annual leave"])
        self.leave_type_combo.grid(row=1, column=1, padx=5, pady=5)
        self.leave_type_combo.current(0)       
        label_start_date = tk.Label(self.popup, text="Start Date (DD-MM-YYYY):")
        label_start_date.grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.calendar_start_date = Calendar(self.popup, selectmode="day", date_pattern="dd-MM-yyyy")
        self.calendar_start_date.grid(row=2, column=1, padx=5, pady=5)      
        label_end_date = tk.Label(self.popup, text="End Date (DD-MM-YYYY):")
        label_end_date.grid(row=3, column=0, padx=5, pady=5, sticky="e")
        self.calendar_end_date = Calendar(self.popup, selectmode="day", date_pattern="dd-MM-yyyy")
        self.calendar_end_date.grid(row=3, column=1, padx=5, pady=5)        
        label_description = tk.Label(self.popup, text="Description:")
        label_description.grid(row=4, column=0, padx=5, pady=5, sticky="e")
        self.entry_description = tk.Text(self.popup, height=5, width=30)
        self.entry_description.grid(row=4, column=1, padx=5, pady=5)       
        save_button = tk.Button(self.popup, text="Save", command=self.validate_and_save_leave_record)
        save_button.grid(row=5, columnspan=2, padx=5, pady=10)

    def validate_and_save_leave_record(self):
        leave_type = self.leave_type_combo.get()
        start_date_str = self.calendar_start_date.get_date()
        end_date_str = self.calendar_end_date.get_date()
        description = self.entry_description.get("1.0", "end-1c")
        
        if not (leave_type and start_date_str and end_date_str):
            messagebox.showerror("Error", "Please fill in all required fields.")
            return        
        try:
            start_date = datetime.strptime(start_date_str, "%d-%m-%Y")
            end_date = datetime.strptime(end_date_str, "%d-%m-%Y")
            
            if start_date > end_date:
                messagebox.showerror("Error", "End date cannot be before start date.")
                return            
            month = start_date.month
            year = start_date.year
            day_difference = (end_date - start_date).days + 1
            conn = sqlite3.connect("employees.db")
            cursor = conn.cursor()
            if leave_type == "Sick leave":
                cursor.execute("""
                    INSERT INTO leave_records 
                    (Employee_ID, Employee_Name, Year, Month, Sick_leave, Start_Date, End_Date, Description) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (self.employee_id, self.employee_name, year, month, day_difference, start_date_str, end_date_str, description))
            elif leave_type == "Annual leave":
                cursor.execute("""
                    INSERT INTO leave_records 
                    (Employee_ID, Employee_Name, Year, Month, Annual_leave, Start_Date, End_Date, Description) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (self.employee_id, self.employee_name, year, month, day_difference, start_date_str, end_date_str, description))           
            conn.commit()
            conn.close()           
            messagebox.showinfo("Success", "Leave record saved successfully!")
            self.popup.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving leave record: {e}")

class EmployeeManagementApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Employee List Management")
        if not os.path.exists("employees.db"):
            self.create_database_with_triggers("employees.db")
        self.conn = sqlite3.connect("employees.db")
        self.create_widgets()
        self.load_employees()

    def create_widgets(self):
        self.tree = ttk.Treeview(self.root, columns=("ID", "Name", "Position", "Status"), show="headings")
        self.tree.heading("ID", text="ID")
        self.tree.heading("Name", text="Name")
        self.tree.heading("Position", text="Position")
        self.tree.heading("Status", text="Status")
        self.tree.pack(padx=10, pady=10)
        filter_frame = tk.Frame(self.root)
        filter_frame.pack(padx=10, pady=(0, 10))
        self.filter_entries = {}
        for idx, column in enumerate(["ID", "Name", "Position", "Status"]):
            filter_label = tk.Label(filter_frame, text=column)
            filter_label.grid(row=0, column=idx*2)
            self.filter_entries[column] = tk.Entry(filter_frame)
            self.filter_entries[column].grid(row=0, column=idx*2+1)
        filter_button = tk.Button(filter_frame, text="Filter", command=self.filter_employees)
        filter_button.grid(row=0, column=len(self.filter_entries)*2, padx=5)
        add_work_hours_button = tk.Button(self.root, text="Add Work Hours", command=self.open_add_work_hours_popup)
        add_work_hours_button.pack(side="top", padx=10, pady=10)
        add_leave_record_button = tk.Button(self.root, text="Add Leave Record", command=self.open_add_leave_record_popup)
        add_leave_record_button.pack(side="top", padx=10, pady=10)
        show_summary_button = tk.Button(self.root, text="Show Summary", command=self.show_summary)
        show_summary_button.pack(side="top", padx=10, pady=10)
        add_employee_button = tk.Button(self.root, text="Add Employee", command=self.add_employee)
        add_employee_button.pack(side="top", padx=10, pady=10)
        remove_employee_button = tk.Button(self.root, text="Remove Employee", command=self.remove_employee)
        remove_employee_button.pack(side="top", padx=10, pady=10)

    def create_database_with_triggers(self, database_name):
        try:
            conn = sqlite3.connect(database_name)
            cursor = conn.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS employees (
                    ID INTEGER PRIMARY KEY,
                    Name TEXT NOT NULL,
                    Position TEXT,
                    Status TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS work_hours (
                    ID INTEGER PRIMARY KEY,
                    Employee_ID INTEGER NOT NULL,
                    Date TEXT NOT NULL,
                    Start_Time TEXT,
                    End_Time TEXT,
                    Break_Start TEXT,
                    Break_End TEXT,
                    FOREIGN KEY(Employee_ID) REFERENCES employees(ID)
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS leave_records (
                    ID INTEGER PRIMARY KEY,
                    Employee_ID INTEGER NOT NULL,
                    Employee_Name TEXT NOT NULL,
                    Year INTEGER NOT NULL,
                    Month INTEGER NOT NULL,
                    Sick_leave INTEGER DEFAULT 0,
                    Annual_leave INTEGER DEFAULT 0,
                    Start_Date TEXT NOT NULL,
                    End_Date TEXT NOT NULL,
                    Description TEXT,
                    FOREIGN KEY(Employee_ID) REFERENCES employees(ID)
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS "summary" (
                    "Employee_Name" TEXT NULL,
                    "Position" TEXT NULL,
                    "Month" UNKNOWN NULL,
                    "Year" UNKNOWN NULL,
                    "Total_Work_Time" INTEGER NULL,
                    "Total_Break_Time" INTEGER NULL,
                    "Net_Work_Time" INTEGER NULL
                , Total_Annual_Leave INTEGER DEFAULT 0, Total_Sick_Leave INTEGER DEFAULT 0, existing_row_count INTEGER DEFAULT 0, "Employee_ID" INTEGER NULL)
            """)
            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS calculate_leave_totals
                AFTER INSERT ON leave_records
                FOR EACH ROW
                BEGIN
                    UPDATE summary
                    SET Total_Annual_Leave = Total_Annual_Leave + COALESCE(NEW.Annual_leave, 0),
                        Total_Sick_Leave = (
                            SELECT SUM(Sick_leave) FROM leave_records WHERE Employee_ID = NEW.Employee_ID AND Month = NEW.Month AND Year = NEW.Year
                        )
                    WHERE Employee_ID = NEW.Employee_ID AND Month = NEW.Month AND Year = NEW.Year;
                END;
                """)
 
            cursor.execute("""
                    CREATE TRIGGER IF NOT EXISTS update_summary
                    AFTER INSERT ON work_hours
                    BEGIN
                        UPDATE summary 
                        SET existing_row_count = (
                            SELECT COUNT(*) FROM summary 
                            WHERE Employee_Name = (SELECT Name FROM employees WHERE ID = NEW.Employee_ID)
                            AND Month = SUBSTR(NEW.Date, 4, 2)
                            AND Year = SUBSTR(NEW.Date, 7, 4)
                        );
                        UPDATE summary 
                        SET 
                            Total_Work_Time = Total_Work_Time + 
                                CASE 
                                    WHEN (SELECT existing_row_count FROM summary) > 0 THEN (STRFTIME('%s', NEW.End_Time) - STRFTIME('%s', NEW.Start_Time))
                                    ELSE 0
                                END,
                            Total_Break_Time = Total_Break_Time + 
                                CASE 
                                    WHEN (SELECT existing_row_count FROM summary) > 0 THEN IFNULL((STRFTIME('%s', NEW.Break_End) - STRFTIME('%s', NEW.Break_Start)), 0)
                                    ELSE 0
                                END,
                            Net_Work_Time = Net_Work_Time + 
                                CASE 
                                    WHEN (SELECT existing_row_count FROM summary) > 0 THEN ((STRFTIME('%s', NEW.End_Time) - STRFTIME('%s', NEW.Start_Time)) - IFNULL((STRFTIME('%s', NEW.Break_End) - STRFTIME('%s', NEW.Break_Start)), 0))
                                    ELSE ((STRFTIME('%s', NEW.End_Time) - STRFTIME('%s', NEW.Start_Time)) - IFNULL((STRFTIME('%s', NEW.Break_End) - STRFTIME('%s', NEW.Break_Start)), 0))
                                END
                        WHERE 
                            Employee_ID = NEW.Employee_ID
                            AND Month = SUBSTR(NEW.Date, 4, 2)
                            AND Year = SUBSTR(NEW.Date, 7, 4);
                        INSERT INTO summary (
                            Employee_Name,
                            Position,
                            Month,
                            Year,
                            Total_Work_Time,
                            Total_Break_Time,
                            Net_Work_Time,
                            Employee_ID
                        ) 
                        SELECT 
                            (SELECT Name FROM employees WHERE ID = NEW.Employee_ID),
                            (SELECT Position FROM employees WHERE ID = NEW.Employee_ID),
                            SUBSTR(NEW.Date, 4, 2),
                            SUBSTR(NEW.Date, 7, 4),
                            (STRFTIME('%s', NEW.End_Time) - STRFTIME('%s', NEW.Start_Time)),
                            IFNULL((STRFTIME('%s', NEW.Break_End) - STRFTIME('%s', NEW.Break_Start)), 0),
                            (STRFTIME('%s', NEW.End_Time) - STRFTIME('%s', NEW.Start_Time)) - IFNULL((STRFTIME('%s', NEW.Break_End) - STRFTIME('%s', NEW.Break_Start)), 0),
                            NEW.Employee_ID
                        WHERE NOT EXISTS (
                            SELECT 1 FROM summary 
                            WHERE Employee_Name = (SELECT Name FROM employees WHERE ID = NEW.Employee_ID)
                            AND Month = SUBSTR(NEW.Date, 4, 2)
                            AND Year = SUBSTR(NEW.Date, 7, 4)
                        );
                    END
                           """)
            conn.commit()
            conn.close()
            print("The database and triggers have been created successfully.")
        except Exception as e:
            print("An error occurred while creating the database and triggers:", e)
       
    def open_add_leave_record_popup(self):
        selected_item = self.tree.focus()
        if selected_item:
            employee_id = self.tree.item(selected_item, "values")[0]
            employee_name = self.tree.item(selected_item, "values")[1]
            AddLeaveRecordPopup(self.root, employee_id=employee_id, employee_name=employee_name)
        else:
            messagebox.showerror("Error", "Please select an employee.")

    def show_summary(self):
        SummaryPopup(self.root)

    def remove_employee(self):
        selected_item = self.tree.focus()
        if selected_item:
            employee_id = self.tree.item(selected_item, "values")[0]
            employee_name = self.tree.item(selected_item, "values")[1]
            confirmation = messagebox.askyesno("Confirmation", f"Are you sure you want to remove {employee_name}?")
            if confirmation:
                try:
                    conn = sqlite3.connect("employees.db")
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM employees WHERE ID=?", (employee_id,))
                    conn.commit()
                    conn.close()
                    self.refresh_treeview()
                    print(f"{employee_name} removed successfully!")
                except Exception as e:
                    print("An error occurred while removing employee:", e)
        else:
            messagebox.showerror("Error", "Please select an employee.")

    def add_employee(self):
        name = self.filter_entries["Name"].get()
        position = self.filter_entries["Position"].get()
        status = self.filter_entries["Status"].get()
        try:
            if not name:
                messagebox.showerror("Error", "Employee name cannot be empty!")
                return
            user_input = messagebox.askyesno("Confirmation", f"Employee Name: {name}\nEmployee Position: {position}\nEmployee Status: {status}\n\nDo you want to save this employee?")
            if user_input:
                conn = sqlite3.connect("employees.db")
                cursor = conn.cursor()
                cursor.execute("SELECT MAX(ID) FROM employees")
                result = cursor.fetchone()
                if result[0] is None:
                    new_id = 1
                else:
                    new_id = result[0] + 1
                cursor.execute("INSERT INTO employees (ID, Name, Position, Status) VALUES (?, ?, ?, ?)", (new_id, name, position, status))
                conn.commit()
                conn.close()
                self.refresh_treeview()
                messagebox.showinfo("Success", "Employee added successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while adding employee: {e}")


    def open_add_work_hours_popup(self):
        selected_item = self.tree.focus()
        if selected_item:
            employee_id = self.tree.item(selected_item, "values")[0]
            employee_name = self.tree.item(selected_item, "values")[1]
            AddWorkHoursPopup(self.root, self.save_work_hours, employee_id=employee_id, employee_name=employee_name)
        else:
            messagebox.showerror("Error", "Please select an employee.")

    def load_employees(self):
        try:
            conn = sqlite3.connect("employees.db")
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM employees")
            employees = cursor.fetchall()
            conn.close()
            for employee in employees:
                self.tree.insert("", "end", values=employee)
        except Exception as e:
            print("An error occurred while loading employees:", e)

    def save_work_hours(self, employee_id, employee_name, work_hours):
        pass

    def refresh_treeview(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        try:
            conn = sqlite3.connect("employees.db")
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM employees")
            employees = cursor.fetchall()
            conn.close()
            for employee in employees:
                self.tree.insert("", "end", values=employee)
        except Exception as e:
            print("An error occurred while fetching employees:", e)

    def filter_employees(self):
        filters = {}
        for column in self.filter_entries:
            value = self.filter_entries[column].get()
            if value:
                filters[column] = value
        try:
            conn = sqlite3.connect("employees.db")
            cursor = conn.cursor()
            conditions = []
            values = ()
            for column, value in filters.items():
                conditions.append(f"{column} LIKE ?")
                values += (f"%{value}%",)
            query = "SELECT * FROM employees"
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
            cursor.execute(query, values)
            filtered_employees = cursor.fetchall()
            conn.close()
            for item in self.tree.get_children():
                self.tree.delete(item)
            for employee in filtered_employees:
                self.tree.insert("", "end", values=employee)
        except Exception as e:
            print("An error occurred while filtering employees:", e)

if __name__ == "__main__":
    root = tk.Tk()
    app = EmployeeManagementApp(root)
    root.mainloop()
